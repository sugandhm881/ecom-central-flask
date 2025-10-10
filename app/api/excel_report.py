from flask import Blueprint, request, Response, current_app
from ..auth import token_required
from .helpers import get_facebook_ads, get_order_source_term, normalize_status, pick_date_for_filter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
from datetime import datetime
import traceback
import os
import json

excel_report_bp = Blueprint('excel_report', __name__)
MASTER_DATA_FILE = 'master_order_data.json'

@excel_report_bp.route('/download-excel-report', methods=['GET'])
@token_required
def download_excel_report():
    since = request.args.get('since')
    until = request.args.get('until')
    date_filter_type = request.args.get('date_filter_type', 'order_date')
    config = current_app.config
    start_date = datetime.strptime(since, '%Y-%m-%d').date()
    end_date = datetime.strptime(until, '%Y-%m-%d').date()

    try:
        print(f"\n--- [Excel Report] Loading data | filter={date_filter_type} ---")
        if not os.path.exists(MASTER_DATA_FILE):
            return "Master data file not found. Please run data_fetcher.py first.", 500
        
        with open(MASTER_DATA_FILE, 'r') as f:
            all_orders = json.load(f)

        shopify_orders_in_range = []
        for o in all_orders:
            d = pick_date_for_filter(o, date_filter_type)
            if d and start_date <= d <= end_date:
                shopify_orders_in_range.append(o)
        
        print(f"Filtered to {len(shopify_orders_in_range)} orders for Excel export")
        
        fb_ads = get_facebook_ads(config, since, until)
        fb_ad_map = {ad['ad_id']: ad for ad in fb_ads}
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Detailed Order Report"

        headers = [
            "Order ID", "Order Date", "Shipped Date", "Delivered Date",
            "Order Amount", "Normalized Status", "Raw Shipment Status",
            "AWB Number", "Courier", "Customer Name", "Email", "Phone",
            "City", "State", "Pincode", "Products (SKU x Qty)",
            "Attribution Source", "UTM Term", "Ad Set Name", "Ad Name", "Campaign Name"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for order in shopify_orders_in_range:
            source, term = get_order_source_term(order)
            raw_status = order.get('raw_rapidshyp_status', order.get('fulfillment_status') or 'Unfulfilled')
            status = normalize_status(order, raw_status)
            
            ad_set_name, ad_name, campaign_name = "N/A", "N/A", "N/A"
            if source == 'facebook_ad':
                matched_ad = fb_ad_map.get(term)
                if matched_ad:
                    ad_set_name = matched_ad.get('adset_name', 'N/A')
                    ad_name = matched_ad.get('ad_name', 'N/A')
                    campaign_name = matched_ad.get('campaign_name', 'N/A')

            shipping_address = order.get('shipping_address', {}) or {}
            awb = order.get('awb')
            courier = next((f.get('tracking_company') for f in order.get('fulfillments', []) if f.get('tracking_company')), None)
            products_str = ", ".join([f"{item.get('sku', 'N/A')} x {item.get('quantity', 0)}" for item in order.get('line_items', [])])

            # Format dates safely
            def format_date(dt_str):
                if not dt_str:
                    return 'N/A'
                try:
                    return datetime.fromisoformat(dt_str.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
                except:
                    return dt_str

            order_date = format_date(order.get('created_at'))
            shipped_date = format_date(order.get('shipped_at'))
            delivered_date = format_date(order.get('delivered_at'))

            row_data = [
                order.get('name'), order_date, shipped_date, delivered_date,
                float(order.get('total_price', 0)), status, raw_status,
                awb, courier,
                f"{shipping_address.get('first_name', '')} {shipping_address.get('last_name', '')}".strip(),
                order.get('email'), shipping_address.get('phone'), shipping_address.get('city'),
                shipping_address.get('province'), shipping_address.get('zip'), products_str,
                source if source != 'facebook_ad' else 'Facebook Ad', term,
                ad_set_name, ad_name, campaign_name
            ]
            ws.append(row_data)
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        
        return Response(
            virtual_workbook,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment;filename=detailed_report_{since}_to_{until}.xlsx'}
        )
    except Exception as e:
        print(f"--- [CRITICAL Excel Report ERROR] ---")
        traceback.print_exc()
        return "An error occurred during Excel report generation.", 500